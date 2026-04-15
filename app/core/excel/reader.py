# app/core/excel/reader.py
from dataclasses import dataclass, field
from typing import List
from pathlib import Path
import openpyxl


@dataclass
class Learner:
    klasse: str
    nachname: str
    vorname: str
    schueler_id: str = ''
    row: int = 0
    is_new: bool = False
    photographed: bool = False  # True wenn bereits in der Excel-Datei als fotografiert markiert


class ExcelReader:
    def __init__(self, path: Path, mapping: dict):
        self.path = path
        self.mapping = mapping
        if not path.exists():
            raise IOError(f'Datei nicht gefunden: {path}')
        try:
            self.wb = openpyxl.load_workbook(path)
        except Exception as e:
            raise IOError(f'Konnte Excel-Datei nicht laden: {e}')

    def locations(self) -> List[str]:
        return self.wb.sheetnames

    def classes_for_location(self, location: str) -> List[str]:
        sheet = self.wb[location]
        col = self.mapping['klasse']
        values = {str(cell.value) for cell in sheet[col][1:] if cell.value}
        return sorted(values)

    def learners(
        self,
        location: str,
        class_name: str,
        skip_photographed: bool = False,
    ) -> List[Learner]:
        """Return all learners for *class_name* in *location*.

        If *skip_photographed* is True, students already marked as
        photographed (Spalte "fotografiert" == "Ja") are excluded.
        """
        sheet = self.wb[location]
        m = self.mapping
        fotografiert_col = m.get('fotografiert')
        result = []
        for row in sheet.iter_rows(min_row=2):
            if str(row[openpyxl.utils.column_index_from_string(m['klasse'])-1].value) == class_name:
                nachname = row[openpyxl.utils.column_index_from_string(m['nachname'])-1].value
                vorname = row[openpyxl.utils.column_index_from_string(m['vorname'])-1].value
                sid = row[openpyxl.utils.column_index_from_string(m['schuelerId'])-1].value
                if nachname and vorname and sid:
                    photographed = False
                    if fotografiert_col:
                        phot_val = row[openpyxl.utils.column_index_from_string(fotografiert_col)-1].value
                        photographed = str(phot_val or '').strip().lower() == 'ja'
                    if skip_photographed and photographed:
                        continue
                    result.append(Learner(
                        str(class_name),
                        str(nachname),
                        str(vorname),
                        str(sid),
                        row=row[0].row,
                        photographed=photographed,
                    ))
        result.sort(key=lambda l: (l.nachname, l.vorname))
        return result

    def duplicate_ids(self, location: str, class_name: str) -> List[str]:
        """Return a list of student IDs that appear more than once in the class."""
        all_learners = self.learners(location, class_name)
        seen: set = set()
        dupes: List[str] = []
        for learner in all_learners:
            sid = learner.schueler_id
            if sid:
                if sid in seen and sid not in dupes:
                    dupes.append(sid)
                seen.add(sid)
        return dupes

    def mark_photographed(
        self,
        location: str,
        row: int,
        photographed: bool,
        date: str | None = None,
        reason: str | None = None,
    ) -> None:
        sheet = self.wb[location]
        col_phot = self.mapping.get('fotografiert')
        col_date = self.mapping.get('aufnahmedatum')
        col_reason = self.mapping.get('grund')
        if col_phot:
            sheet[f"{col_phot}{row}"].value = 'Ja' if photographed else 'Nein'
        if col_date:
            sheet[f"{col_date}{row}"].value = date if photographed else None
        if col_reason:
            sheet[f"{col_reason}{row}"].value = None if photographed else reason
        try:
            self.wb.save(self.path)
        except Exception as e:
            raise IOError(f'Konnte Excel-Datei nicht speichern: {e}')
