# app/core/excel/test_data.py
"""Generate a randomized placeholder Excel roster for one-shot Test Mode.

Builds a workbook structurally identical to a real school roster (Standort
sheets with Klasse/Nachname/Vorname/SchuelerID columns) using the app's
existing ``excelMapping`` column-letter scheme, so ``ExcelReader`` can load it
exactly like a real file. Photo/date/reason cells are left blank, matching a
fresh, not-yet-photographed roster.
"""
import random
from pathlib import Path

import openpyxl

FIRST_NAMES = [
    "Anna", "Lena", "Mia", "Noah", "Luca", "Elias", "Sara", "Tim", "Nina",
    "Jonas", "Emma", "Leon", "Julia", "Finn", "Laura", "David", "Sophie",
    "Ben", "Lea", "Marco",
]
LAST_NAMES = [
    "Meier", "Keller", "Fischer", "Weber", "Huber", "Steiner", "Baumann",
    "Brunner", "Frei", "Graf", "Hofer", "Kaufmann", "Lehmann", "Moser",
    "Roth", "Schmid", "Suter", "Vogel", "Wyss", "Zimmermann",
]


def generate_test_roster(
    path: Path,
    mapping: dict,
    locations: int = 2,
    classes_per_location: int = 3,
    students_per_class: tuple[int, int] = (6, 10),
    seed: int | None = None,
) -> None:
    """Write a randomized placeholder roster to *path*.

    *mapping* is a dict with keys ``klasse``/``nachname``/``vorname``/
    ``schuelerId``/``fotografiert``/``aufnahmedatum``/``grund`` mapping each
    logical column to an Excel column letter (same shape as
    ``settings.excelMapping.model_dump()``).
    """
    rng = random.Random(seed)
    pool = [(v, n) for v in FIRST_NAMES for n in LAST_NAMES]
    rng.shuffle(pool)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = {
        mapping['klasse']: 'Klasse',
        mapping['nachname']: 'Nachname',
        mapping['vorname']: 'Vorname',
        mapping['schuelerId']: 'SchuelerID',
        mapping['fotografiert']: 'Fotografiert?',
        mapping['aufnahmedatum']: 'Aufnahmedatum',
        mapping['grund']: 'Grund',
    }

    student_no = 1
    for loc_idx in range(1, locations + 1):
        ws = wb.create_sheet(title=f"Testort {loc_idx}")
        for col, label in headers.items():
            ws[f"{col}1"] = label
        row = 2
        for cls_idx in range(1, classes_per_location + 1):
            class_name = f"T{loc_idx}{cls_idx}"
            n = rng.randint(*students_per_class)
            for _ in range(n):
                vorname, nachname = pool.pop()
                ws[f"{mapping['klasse']}{row}"] = class_name
                ws[f"{mapping['nachname']}{row}"] = nachname
                ws[f"{mapping['vorname']}{row}"] = vorname
                ws[f"{mapping['schuelerId']}{row}"] = f"T{student_no:04d}"
                student_no += 1
                row += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
