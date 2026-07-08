# app/core/excel/missed_writer.py
from dataclasses import dataclass
from pathlib import Path
import openpyxl
from datetime import datetime

@dataclass
class MissedEntry:
    standort: str
    klasse: str
    nachname: str
    vorname: str
    schueler_id: str
    datum: str
    grund: str = ''

class MissedWriter:
    HEADER = ['Standort', 'Klasse', 'Nachname', 'Vorname', 'SchuelerID', 'Datum', 'Grund']

    def __init__(self, path: Path):
        self.path = path
        if path.exists():
            self.wb = openpyxl.load_workbook(path)
            self.ws = self.wb.active
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(self.HEADER)
            self.wb.save(path)

    def append(self, entry: MissedEntry) -> None:
        try:
            # Update an existing row for the same person instead of adding a
            # duplicate if they were skipped more than once in a session.
            existing_row = None
            for row in self.ws.iter_rows(min_row=2):
                standort, klasse, _nachname, _vorname, schueler_id = (
                    row[0].value, row[1].value, row[2].value, row[3].value, row[4].value,
                )
                if (standort, klasse, schueler_id) == (entry.standort, entry.klasse, entry.schueler_id):
                    existing_row = row
                    break
            if existing_row is not None:
                existing_row[5].value = entry.datum
                existing_row[6].value = entry.grund
            else:
                self.ws.append([
                    entry.standort,
                    entry.klasse,
                    entry.nachname,
                    entry.vorname,
                    entry.schueler_id,
                    entry.datum,
                    entry.grund,
                ])
            self.wb.save(self.path)
        except Exception as e:
            raise IOError(f'Konnte Datei für verpasste Termine nicht speichern: {e}')
